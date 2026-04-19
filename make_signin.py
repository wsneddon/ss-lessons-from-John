from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()
ws = wb.active
ws.title = "Sign-In"

# Title
ws.merge_cells("A1:E1")
ws["A1"] = "Select Teachings from the Gospel of John"
ws["A1"].font = Font(bold=True, size=14, color="4A7099")
ws["A1"].alignment = Alignment(horizontal="center")

ws.merge_cells("A2:E2")
ws["A2"] = "Date: "
ws["A2"].font = Font(size=11)

# Headers
headers = ["#", "Name", "Email", "Phone", "Receive Materials?"]
header_fill = PatternFill("solid", fgColor="4A7099")
header_font = Font(bold=True, color="FFFFFF")
thin = Side(style="thin", color="CCCCCC")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for col, h in enumerate(headers, 1):
    cell = ws.cell(row=3, column=col, value=h)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center" if col in (1, 5) else "left")
    cell.border = border

# Column widths
ws.column_dimensions["A"].width = 5
ws.column_dimensions["B"].width = 25
ws.column_dimensions["C"].width = 30
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18

# Data rows
for row in range(4, 24):
    ws.cell(row=row, column=1, value=row - 3).alignment = Alignment(horizontal="center")
    for col in range(1, 6):
        cell = ws.cell(row=row, column=col)
        cell.border = border
        cell.alignment = Alignment(horizontal="center" if col in (1, 5) else "left", vertical="center")
    ws.cell(row=row, column=5, value="☐").alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 20

wb.save("signin-sheet.xlsx")
print("Created signin-sheet.xlsx")
